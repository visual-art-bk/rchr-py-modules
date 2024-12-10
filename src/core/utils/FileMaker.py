import os
import json
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from app.core.utils.Logger import Logger

logger = Logger(name="FileMaker", log_file="FileMaker.log").get_logger()


class FileMaker:
    def __init__():
        pass

    @staticmethod
    def save_list_to_json(list):
        if len(list) == 0:
            logger.info("json파일저장실패, 저장된 링크 없음")
            return False

        json_file_path = ".data/musinsa_event_links.json"

        with open(json_file_path, "w", encoding="utf-8") as json_file:
            json.dump(list, json_file, ensure_ascii=False, indent=4)

        print(f"총 {len(list)}개의 링크를 {json_file_path} 파일에 저장완료")

    @staticmethod
    def save_to_excel_for_musinsa(
        infos_list,
        file_name="infos_list",
        fixed_columns=None,
        root_dir="상표권출원등록사진",
    ):
        if fixed_columns is None:
            fixed_columns = []

        # infos_list를 데이터프레임으로 변환
        df = pd.DataFrame(infos_list)

        # 동적으로 추가되는 칼럼 추출
        dynamic_columns = [col for col in df.columns if col not in fixed_columns]

        # 최종 칼럼 순서: 고정 칼럼 + 동적 칼럼
        final_columns = fixed_columns + dynamic_columns

        # 누락된 칼럼은 빈 문자열로 채우기
        for col in final_columns:
            if col not in df.columns:
                df[col] = ""

        # 데이터프레임을 최종 칼럼 순서로 정렬
        df = df[final_columns]

        # 임시 엑셀 파일로 저장
        temp_file = f"{file_name}_temp.xlsx"
        df.to_excel(temp_file, index=False)

        # openpyxl을 사용하여 하이퍼링크 추가 및 사진 삽입
        workbook = load_workbook(temp_file)
        sheet = workbook.active

        # "브랜드 페이지" 열의 위치를 찾기
        brand_page_column = None
        for col in sheet.iter_cols(1, sheet.max_column, 1, 1):
            if col[0].value == "브랜드 페이지":
                brand_page_column = col[0].column_letter
                break

        # "브랜드 페이지" 열에 하이퍼링크 추가
        for row in range(2, sheet.max_row + 1):  # 헤더 제외
            cell = sheet[f"{brand_page_column}{row}"]
            url = cell.value
            if url:
                cell.value = "바로가기"
                cell.hyperlink = url
                cell.font = Font(color="0000FF", underline="single")

        # Fixed columns 셀 값 세로 기준 가운데 정렬
        for col in fixed_columns:
            if col in final_columns:
                col_idx = final_columns.index(col) + 1  # 1부터 시작하는 엑셀 인덱스
                col_letter = get_column_letter(col_idx)
                for row in range(2, sheet.max_row + 1):  # 헤더 제외
                    cell = sheet[f"{col_letter}{row}"]
                    cell.alignment = Alignment(horizontal="center", vertical="center")


        # 동적으로 추가된 칼럼에 사진 삽입 및 정렬
        for col in dynamic_columns:
            col_idx = final_columns.index(col) + 1  # 엑셀 컬럼 인덱스는 1부터 시작
            col_letter = get_column_letter(col_idx)

            for row in range(2, sheet.max_row + 1):  # 헤더 제외
                brand = sheet[f"A{row}"].value  # 브랜드 셀 값 가져오기 (A열 가정)
                application_number = sheet[f"{col_letter}{row}"].value  # 출원번호 값 가져오기

                if not (brand and application_number):  # 값이 없으면 건너뜀
                    continue

                # 루트 디렉토리에서 브랜드 하위 디렉토리 찾기
                brand_dir = os.path.join(root_dir, brand)
                if not os.path.isdir(brand_dir):
                    continue

                # 하위 디렉토리에서 출원번호와 일치하는 jpg 파일 찾기
                image_path = os.path.join(brand_dir, f"{application_number}.jpg")
                if os.path.isfile(image_path):
                    # 셀 값 유지
                    cell = sheet[f"{col_letter}{row}"]
                    cell.value = application_number

                    # 텍스트 정렬
                    cell.alignment = Alignment(horizontal="center", vertical="bottom")

                    # 사진 삽입
                    img = Image(image_path)
                    img.height = 60  # 이미지 높이 조정
                    img.width = 60  # 이미지 너비 조정
                    img.anchor = f"{col_letter}{row}"
                    sheet.add_image(img)

                    # 행 높이와 열 너비를 이미지 크기에 맞춤
                    sheet.row_dimensions[row].height = img.height + 10
                    sheet.column_dimensions[col_letter].width = (img.width / 7) + 2  # 엑셀의 열 너비 단위 변환

        # 칼럼 너비를 자동으로 조정
        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in column_cells:
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = max_length + 10
            sheet.column_dimensions[col_letter].width = adjusted_width

        # 최종 엑셀 파일 저장
        if not file_name.endswith(".xlsx"):
            final_file = f"{file_name}.xlsx"
        else:
            final_file = file_name

        workbook.save(final_file)
        workbook.close()

        # temp.xlsx 삭제
        try:
            os.remove(temp_file)
            print(f"임시 파일 {temp_file}이 삭제되었습니다.")
        except Exception as e:
            print(f"임시 파일 {temp_file} 삭제 중 오류 발생: {e}")

        print(f"파일이 저장되었습니다: {final_file}")
